"""
Generatore di report Excel per il Bond Scanner.

Crea un file Excel nel formato identico a "Lista acquisti-2026.xlsx":
- 3 fogli: Euro, USD, GBP
- Colonne: PAPY, Nome, Emissione, ISIN, Price, Yield, Rating, Volume, Min. piece
- Colorazione rosso/nero secondo il prezzo
- Ordinamento per yield decrescente

Tutto in italiano 🇮🇹
"""

import logging
import os
from datetime import date, datetime
from typing import List, Dict

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scanner.models import ScannedBond

logger = logging.getLogger(__name__)

# Mesi italiani
ITALIAN_MONTHS = {
    1: 'gen', 2: 'feb', 3: 'mar', 4: 'apr', 5: 'mag', 6: 'giu',
    7: 'lug', 8: 'ago', 9: 'set', 10: 'ott', 11: 'nov', 12: 'dic',
}

# Stili
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="2F5496")
DATA_FONT = Font(name="Calibri", size=10, color="000000")
RED_FONT = Font(name="Calibri", size=10, color="FF0000")
# Lien hypertexte Fitch (fix 2026-05-29) : le rating de la col. G pointe vers
# la page Fitch qui le prouve → Massii clique et vérifie à la main.
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

# Mapping foglio → valuta
SHEET_CURRENCY_MAP = {
    'Euro': 'EUR',
    'USD': 'USD',
    'GBP': 'GBP',
}


def _format_date_italian(d) -> str:
    """Formatta una data in formato italiano: 'mag.25', 'feb.23', etc."""
    if isinstance(d, str):
        for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
            try:
                d = datetime.strptime(d, fmt).date()
                break
            except ValueError:
                continue
        else:
            return d
    if isinstance(d, (date, datetime)):
        month_abbr = ITALIAN_MONTHS.get(d.month, str(d.month))
        year_short = str(d.year)[-2:]
        return f"{month_abbr}.{year_short}"
    return str(d)


def _format_as_k(value) -> str:
    """Formatta un numero in notazione 'Xk': 2000→'2k', 500000000→'500,000k'."""
    try:
        num = float(str(value).replace(',', '').replace("'", ''))
        k_value = int(num / 1000)
        if k_value == 0:
            return str(int(num))
        return f"{k_value:,}k"
    except (ValueError, TypeError):
        return str(value) if value else ""


def _format_bond_name(bond: ScannedBond) -> str:
    """
    Formatta il nome del bond come nel file Lista acquisti:
    "ISSUER COUPON% - DD.MM.YY" oppure "ISSUER COUPON,XXX - DD.MM.YY"
    """
    name = bond.name or ""

    # Se il nome è già nel formato giusto, usarlo
    if ' - ' in name and any(c.isdigit() for c in name):
        return name

    # Costruire il nome dal nome + cedola + scadenza
    parts = [name]

    if bond.coupon_rate is not None:
        coupon_str = f"{bond.coupon_rate}".replace('.', ',')
        if ',' in coupon_str:
            coupon_str = coupon_str.rstrip('0').rstrip(',')
        parts.append(f"{coupon_str}%")

    if bond.maturity_date is not None:
        date_str = bond.maturity_date.strftime('%d.%m.%y')
        parts.append(f"- {date_str}")

    return ' '.join(parts)


class ReportGenerator:
    """
    Generatore di report Excel per le obbligazioni trovate.

    Crea un file identico al formato "Lista acquisti-2026.xlsx"
    con le obbligazioni trovate durante la scansione.
    """

    def __init__(self, price_threshold: float = 101.0):
        """
        Args:
            price_threshold: Soglia di prezzo per la colorazione rosso/nero
        """
        self.price_threshold = price_threshold
        self.wb = openpyxl.Workbook()

    def generate(
        self,
        bonds: List[ScannedBond],
        output_path: str,
        criteria_info: str = "",
        timed_out: bool = False,
        quota_exhausted: bool = False,
        quota_low: bool = False,
    ) -> str:
        """
        Genera il file Excel con le obbligazioni trovate.

        Args:
            bonds: Lista di ScannedBond da includere nel report
            output_path: Percorso del file di output
            criteria_info: Descrizione dei criteri usati (per il titolo)
            timed_out: Se True, prepende un banner ai criteri segnalando
                       che la scansione è stata interrotta dal timeout
                       45 min (Task 18) e che i risultati sono parziali.
            quota_exhausted: Se True, prepende un banner sull'esaurimento
                       della quota Brave Search API (1000 req/mese free
                       dépassées). Lo scan è stato interrotto perché
                       impossibile recuperare altri rating Fitch.

        Returns:
            Percorso del file generato
        """
        if quota_exhausted:
            # Banner prioritario : se la quota è esaurita c'est probablement
            # la cause principale de l'interruption, à signaler en premier.
            criteria_info = (
                "🚫 BRAVE SEARCH API QUOTA ESAURITA — IMPOSSIBILE RECUPERARE "
                "ALTRI RATING FITCH. RISULTATI PARZIALI (solo i bond rated "
                "prima dell'esaurimento + quelli già in cache). Pose une "
                "nouvelle clé via le panel admin OR upgrade le plan Brave. "
                + criteria_info
            )
        if quota_low:
            # Réserve Brave ≤ 50 → arrêt préventif (garde-fou Massii).
            criteria_info = (
                "🛑 RÉSERVE BRAVE ≤ 50 REQUÊTES — SCAN ARRÊTÉ POUR PRÉSERVER "
                "LA QUOTA. RÉSULTATS PARTIELS. Attends le reset mensuel ou "
                "augmente le plan Brave pour scanner plus. "
                + criteria_info
            )
        if timed_out:
            criteria_info = (
                "⏱ SCANSIONE INTERROTTA A 45 MIN — RISULTATI PARZIALI. "
                + criteria_info
            )

        # Raggruppa per valuta
        groups: Dict[str, List[ScannedBond]] = {
            'EUR': [], 'USD': [], 'GBP': [],
        }

        for bond in bonds:
            currency = bond.currency.upper() if bond.currency else 'EUR'
            if currency in groups:
                groups[currency].append(bond)

        # Crea i fogli
        self._remove_default_sheet()

        for sheet_name, currency in SHEET_CURRENCY_MAP.items():
            # Task 15 (2026-05-28) : i bond arrivano già ordinati per
            # composite_score desc (calcolato da scanner.scoring). Se il
            # caller non ha pre-ordinato (legacy), fallback su yield desc.
            sheet_bonds = groups.get(currency, [])
            if sheet_bonds and any(b.composite_score for b in sheet_bonds):
                # Già scored → preserva l'ordine di top_n_per_currency
                pass
            else:
                sheet_bonds = sorted(
                    sheet_bonds,
                    key=lambda b: b.calculated_yield or 0,
                    reverse=True,
                )
            self._create_standard_sheet(sheet_name, currency, sheet_bonds, criteria_info)

        # Salva
        self.wb.properties.creator = 'Bond Scanner — OmenServer'
        self.wb.properties.lastModifiedBy = 'Bond Scanner'
        self.wb.save(output_path)

        logger.info(f"✅ Report salvato: {output_path}")
        return output_path

    def _remove_default_sheet(self):
        """Rimuove il foglio di default creato da openpyxl."""
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

    def _create_standard_sheet(
        self, sheet_name: str, currency: str,
        bonds: List[ScannedBond], criteria_info: str
    ):
        """Crea un foglio standard (Euro, USD, GBP)."""
        ws = self.wb.create_sheet(title=sheet_name)

        # Largeur des colonnes
        col_widths = {'A': 6, 'B': 42, 'C': 12, 'D': 16, 'E': 14, 'F': 10,
                      'G': 22, 'H': 14, 'I': 10}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Riga 1: vuota
        # Riga 2: titolo valuta
        ws.merge_cells('E2:F2')
        ws['E2'] = currency
        ws['E2'].font = TITLE_FONT
        ws['E2'].alignment = CENTER_ALIGN

        # Riga 3: header
        headers = [
            ('A3', 'PAPY'),
            ('B3', f'{date.today().strftime("%d.%m.%y")} Bond (Bond Scanner)'),
            ('C3', 'Emissione'),
            ('D3', 'ISIN'),
            ('E3', f'Price - {currency}'),
            ('F3', 'Yield'),
            ('G3', 'Rating (fonte)'),
            ('H3', 'Volume'),
            ('I3', 'Min. pie'),
        ]

        for cell_ref, value in headers:
            cell = ws[cell_ref]
            cell.value = value
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Righe dati (a partire da riga 4)
        for i, bond in enumerate(bonds, start=4):
            self._write_bond_row(ws, i, bond, is_vale=False)

        logger.info(f"  📄 Foglio {sheet_name}: {len(bonds)} obbligazioni")

    def _write_bond_row(self, ws, row: int, bond: ScannedBond, is_vale: bool=False):
        """Scrive una riga di dati per un'obbligazione."""
        price = bond.current_price or 0
        font = RED_FONT if price > self.price_threshold else DATA_FONT

        ws[f'A{row}'].alignment = CENTER_ALIGN
        ws[f'A{row}'].border = THIN_BORDER

        ws[f'B{row}'] = _format_bond_name(bond)
        ws[f'B{row}'].font = font
        ws[f'B{row}'].alignment = LEFT_ALIGN
        ws[f'B{row}'].border = THIN_BORDER

        if bond.issue_date:
            ws[f'C{row}'] = _format_date_italian(bond.issue_date)
        ws[f'C{row}'].font = font
        ws[f'C{row}'].alignment = CENTER_ALIGN
        ws[f'C{row}'].border = THIN_BORDER

        ws[f'D{row}'] = bond.isin
        ws[f'D{row}'].font = font
        ws[f'D{row}'].alignment = CENTER_ALIGN
        ws[f'D{row}'].border = THIN_BORDER

        if bond.current_price is not None:
            ws[f'E{row}'] = bond.current_price
        ws[f'E{row}'].font = font
        ws[f'E{row}'].alignment = CENTER_ALIGN
        ws[f'E{row}'].border = THIN_BORDER

        if bond.calculated_yield is not None:
            ws[f'F{row}'] = round(bond.calculated_yield, 4)
            ws[f'F{row}'].number_format = '0.00%'
        ws[f'F{row}'].font = font
        ws[f'F{row}'].alignment = CENTER_ALIGN
        ws[f'F{row}'].border = THIN_BORDER

        # Rating Fitch VÉRIFIÉ (politique fitch_only) : un bond sans rating ne
        # devrait jamais arriver ici (criteria.matches le rejette). Quand on a
        # l'URL Fitch source, on en fait un LIEN cliquable → Massii ouvre la
        # page Fitch qui prouve le rating (demande 2026-05-29 : "un vrai que je
        # peux chercher à la main").
        rating_text = bond.rating_display or bond.rating or ''
        ws[f'G{row}'] = rating_text
        if rating_text and getattr(bond, 'rating_url', None):
            ws[f'G{row}'].hyperlink = bond.rating_url
            ws[f'G{row}'].font = LINK_FONT
        else:
            ws[f'G{row}'].font = font
        ws[f'G{row}'].alignment = CENTER_ALIGN
        ws[f'G{row}'].border = THIN_BORDER

        ws[f'H{row}'] = _format_as_k(bond.volume) if bond.volume else ''
        ws[f'H{row}'].font = font
        ws[f'H{row}'].alignment = CENTER_ALIGN
        ws[f'H{row}'].border = THIN_BORDER

        ws[f'I{row}'] = _format_as_k(bond.min_piece) if bond.min_piece else ''
        ws[f'I{row}'].font = font
        ws[f'I{row}'].alignment = CENTER_ALIGN
        ws[f'I{row}'].border = THIN_BORDER

