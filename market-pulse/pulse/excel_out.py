"""Export Excel du Market Pulse — 3 feuilles : Mercati, Statistiche, Notizie.

Signature visuelle reprise du Bond Scanner (`bond-scanner/excel/report_generator.py`) :
en-têtes bleu 2F5496 sur fond plein, police Calibri, bordures fines gris clair,
titres en italien. Ce module ne l'IMPORTE pas (le générateur du Bond Scanner
dépend de son propre package) — seuls les styles sont recopiés.

Contenu FACTUEL uniquement : cours, variations, gaps, statistiques, titres de
presse. Aucune recommandation, aucune prévision — un avertissement explicite est
posé en tête de la première feuille.

--------------------------------------------------------------------------
CHOIX DE FORMAT NUMÉRIQUE (important, à ne pas « corriger » par inadvertance)
--------------------------------------------------------------------------
Les pourcentages sont stockés TELS QUE PUBLIÉS par le snapshot (change_pct
-0.69 signifie -0,69 %) et affichés via un format à `%` LITTÉRAL entre
guillemets : ``+0.00"%";-0.00"%";0.00"%"``. On n'utilise PAS le format
pourcentage natif d'Excel (`0.00%`) qui exigerait de diviser par 100 à
l'écriture : la valeur en cellule resterait alors 0,0069 et toute lecture
directe (tri, filtre « > 1 », copie vers un autre outil) deviendrait piégeuse
pour l'utilisateur. Avec le choix retenu, ce qu'on lit dans la barre de formule
est exactement le chiffre publié.

Les codes de format sont écrits en notation canonique en-US (point décimal,
virgule de milliers) : c'est Excel qui les restitue selon la locale de la
machine — sur un Excel italien `#,##0.00` s'affiche donc bien « 51.698,19 ».

Cas particulier `kind == "rate"` : le « prix » EST un taux. Le niveau est
affiché en pourcentage (4,59%) et la VARIATION en points de base
(4,592 − 4,641 = −4,9 pb), jamais en pourcentage relatif qui serait trompeur.
"""

from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:  # pragma: no cover - présent en Python 3.9+
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover
    ZoneInfo = None

# Fuseau de lecture de l'utilisateur (investisseur italien).
USER_TZ = "Europe/Rome"

# --- Styles (miroir du Bond Scanner) --------------------------------------
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color="2F5496")
WARNING_FONT = Font(name="Calibri", size=10, bold=True, italic=True, color="C00000")
DATA_FONT = Font(name="Calibri", size=10, color="000000")
GREEN_FONT = Font(name="Calibri", size=10, color="107C10")
RED_FONT = Font(name="Calibri", size=10, color="FF0000")
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
WRAP_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# --- Formats numériques ---------------------------------------------------
FMT_NUM = "#,##0.00"
FMT_FX = "#,##0.0000"
FMT_RATE = '0.00"%"'                                    # niveau d'un taux
FMT_PCT = '+0.00"%";-0.00"%";0.00"%"'                   # variation / gap
FMT_PCT_PLAIN = '0.0"%"'                                # part (% de séances)
FMT_BP = '+0.0" pb";-0.0" pb";0.0" pb"'                 # variation d'un taux
FMT_INT = "#,##0"
FMT_DATE = "DD/MM/YYYY"
FMT_DATETIME = "DD/MM/YYYY HH:MM"

AVVISO = (
    "Documento puramente informativo: dati di mercato e titoli di stampa. "
    "Nessun consiglio di investimento, nessuna raccomandazione di acquisto o "
    "vendita, nessuna previsione."
)

REGION_IT = {
    "europe": "Europa",
    "usa": "USA",
    "asia": "Asia",
    "global": "Globale",
}

KIND_IT = {
    "index": "Indice",
    "future": "Future",
    "fx": "Cambio",
    "commodity": "Materia prima",
    "rate": "Tasso",
    "volatility": "Volatilità",
}

STATUS_IT = {
    "open": "APERTO",
    "closed": "chiuso",
    "unknown": "n.d.",
}

# Ordre des jours tel que produit par `build_history_stats`.
WEEKDAY_ORDER = ["lunedì", "martedì", "mercoledì", "giovedì", "venerdì",
                 "sabato", "domenica"]

MERCATI_HEADERS = [
    "Mercato", "Area", "Tipo", "Stato", "Ora locale", "Ultimo",
    "Chiusura prec.", "Var. %", "Gap apertura %", "Data gap", "Valuta",
]
MERCATI_WIDTHS = [26, 10, 15, 10, 11, 14, 15, 12, 16, 12, 8]

STAT_HEADERS = [
    "Mercato", "Sedute", "Giorno", "Sedute (giorno)",
    "Gap medio %", "Gap medio assoluto %", "Sedute in rialzo %",
]
GAP_HEADERS = ["Mercato", "Data", "Gap %", "Apertura", "Chiusura prec."]
NEWS_HEADERS = ["Titolo", "Fonte", "Ora", "Link"]


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def _safe_text(value: Any) -> Optional[str]:
    """Neutralise l'injection de formule Excel — même posture que
    `harvester_router._csv_safe` : toute chaîne commençant par = + - @ tab ou
    CR est préfixée d'une apostrophe. Empêche aussi openpyxl de typer la
    cellule en formule (`data_type == 'f'`)."""
    if value is None:
        return None
    text = value if isinstance(value, str) else str(value)
    if text and text[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + text
    return text


def _num(value: Any) -> Optional[float]:
    """Nombre exploitable, ou None (jamais la chaîne « None » en cellule)."""
    if value is None or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _parse_date(value: Any):
    """« YYYY-MM-DD » → date. Rend la chaîne telle quelle si non parsable."""
    if not value or not isinstance(value, str):
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _local_dt(epoch: Any):
    """Epoch → datetime NAÏF dans le fuseau de l'utilisateur (openpyxl refuse
    les datetimes porteurs de tzinfo)."""
    ts = _num(epoch)
    if ts is None:
        return None
    try:
        if ZoneInfo is not None:
            return datetime.fromtimestamp(ts, ZoneInfo(USER_TZ)).replace(tzinfo=None)
        return datetime.fromtimestamp(ts)  # pragma: no cover
    except (OverflowError, OSError, ValueError):  # pragma: no cover
        return None


def _sign_font(value: Optional[float]) -> Font:
    """Vert si positif, rouge si négatif, neutre à zéro (ou valeur absente)."""
    if value is None or value == 0:
        return DATA_FONT
    return GREEN_FONT if value > 0 else RED_FONT


def _put(ws, row: int, col: int, value=None, font: Font = DATA_FONT,
         fmt: Optional[str] = None, align: Alignment = CENTER_ALIGN,
         link: Optional[str] = None):
    """Écrit une cellule stylée. `value is None` → cellule VIDE (pas « None »)."""
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = font
    cell.alignment = align
    cell.border = THIN_BORDER
    if fmt and value is not None:
        cell.number_format = fmt
    if link:
        cell.hyperlink = link
        cell.font = LINK_FONT
    return cell


def _write_headers(ws, row: int, headers: List[str]) -> None:
    for idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx)
        cell.value = name
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _set_widths(ws, widths: List[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _variation(kind: Optional[str], price, prev_close, change_pct):
    """(valeur, format) de la variation.

    Pour un taux (`kind == 'rate'`) la lecture correcte est la différence en
    points de base : dire « -1,06 % » d'un rendement qui passe de 4,64 à 4,59
    est trompeur. Pour tout le reste, le pourcentage publié.
    """
    if kind == "rate":
        cur, prev = _num(price), _num(prev_close)
        if cur is not None and prev is not None:
            return round((cur - prev) * 100.0, 1), FMT_BP
    return _num(change_pct), FMT_PCT


def _gap_value(kind: Optional[str], gap: Optional[Dict[str, Any]]):
    """(valeur, format) du gap d'ouverture — en pb pour un taux."""
    if not isinstance(gap, dict):
        return None, FMT_PCT
    if kind == "rate":
        open_, prev = _num(gap.get("open")), _num(gap.get("prev_close"))
        if open_ is not None and prev is not None:
            return round((open_ - prev) * 100.0, 1), FMT_BP
    return _num(gap.get("gap_pct")), FMT_PCT


def _price_format(kind: Optional[str]) -> str:
    if kind == "fx":
        return FMT_FX
    if kind == "rate":
        return FMT_RATE
    return FMT_NUM


# --------------------------------------------------------------------------
# Feuille 1 — Mercati
# --------------------------------------------------------------------------

def _sheet_mercati(wb, snapshot: Dict[str, Any]) -> None:
    ws = wb.create_sheet(title="Mercati")
    _set_widths(ws, MERCATI_WIDTHS)

    generated = _local_dt((snapshot or {}).get("generated_at"))
    title = "Market Pulse — Panoramica dei mercati"
    if generated is not None:
        title += generated.strftime(" (dati del %d/%m/%Y alle %H:%M)")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT

    ws["A2"] = AVVISO
    ws["A2"].font = WARNING_FONT

    header_row = 4
    _write_headers(ws, header_row, MERCATI_HEADERS)

    markets = (snapshot or {}).get("markets") or []
    for offset, market in enumerate(markets):
        _write_market_row(ws, header_row + 1 + offset, market)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if markets:
        last_col = get_column_letter(len(MERCATI_HEADERS))
        ws.auto_filter.ref = "A%d:%s%d" % (
            header_row, last_col, header_row + len(markets))


def _write_market_row(ws, row: int, market: Dict[str, Any]) -> None:
    market = market or {}
    kind = market.get("kind")
    clock = market.get("clock") or {}
    gap = market.get("gap")

    _put(ws, row, 1, _safe_text(market.get("label") or market.get("symbol")),
         align=LEFT_ALIGN)
    _put(ws, row, 2, REGION_IT.get(market.get("region")))
    _put(ws, row, 3, KIND_IT.get(kind))
    _put(ws, row, 4, STATUS_IT.get(clock.get("status")))
    _put(ws, row, 5, _safe_text(clock.get("local_time")))

    price_fmt = _price_format(kind)
    _put(ws, row, 6, _num(market.get("price")), fmt=price_fmt)
    _put(ws, row, 7, _num(market.get("prev_close")), fmt=price_fmt)

    var_value, var_fmt = _variation(
        kind, market.get("price"), market.get("prev_close"),
        market.get("change_pct"))
    _put(ws, row, 8, var_value, font=_sign_font(var_value), fmt=var_fmt)

    gap_value, gap_fmt = _gap_value(kind, gap)
    _put(ws, row, 9, gap_value, font=_sign_font(gap_value), fmt=gap_fmt)

    gap_date = _parse_date(gap.get("date")) if isinstance(gap, dict) else None
    _put(ws, row, 10, gap_date, fmt=FMT_DATE)
    _put(ws, row, 11, _safe_text(market.get("currency")))


# --------------------------------------------------------------------------
# Feuille 2 — Statistiche
# --------------------------------------------------------------------------

def _sheet_statistiche(wb, history: Dict[str, Any]) -> None:
    ws = wb.create_sheet(title="Statistiche")
    _set_widths(ws, [26, 10, 14, 16, 14, 22, 18])

    ws["A1"] = "Statistiche storiche dei gap di apertura (ultimo anno)"
    ws["A1"].font = TITLE_FONT

    stats = (history or {}).get("stats") or {}

    row = 3
    _write_headers(ws, row, STAT_HEADERS)
    row += 1
    for symbol, entry in stats.items():
        entry = entry or {}
        label = _safe_text(entry.get("label") or symbol)
        n_sessions = _num(entry.get("n_sessions"))
        weekdays = entry.get("weekday_stats") or {}
        for day in _ordered_weekdays(weekdays):
            day_stats = weekdays.get(day) or {}
            _put(ws, row, 1, label, align=LEFT_ALIGN)
            _put(ws, row, 2, n_sessions, fmt=FMT_INT)
            _put(ws, row, 3, _safe_text(day), align=LEFT_ALIGN)
            _put(ws, row, 4, _num(day_stats.get("n")), fmt=FMT_INT)
            avg = _num(day_stats.get("avg_gap_pct"))
            _put(ws, row, 5, avg, font=_sign_font(avg), fmt=FMT_PCT)
            _put(ws, row, 6, _num(day_stats.get("avg_abs_gap_pct")), fmt=FMT_PCT_PLAIN)
            _put(ws, row, 7, _num(day_stats.get("pct_up")), fmt=FMT_PCT_PLAIN)
            row += 1

    row += 1
    ws.cell(row=row, column=1).value = "Maggiori gap registrati"
    ws.cell(row=row, column=1).font = SUBTITLE_FONT
    row += 1
    _write_headers(ws, row, GAP_HEADERS)
    row += 1
    for symbol, entry in stats.items():
        entry = entry or {}
        label = _safe_text(entry.get("label") or symbol)
        for gap in (entry.get("biggest_gaps") or []):
            gap = gap or {}
            _put(ws, row, 1, label, align=LEFT_ALIGN)
            _put(ws, row, 2, _parse_date(gap.get("date")), fmt=FMT_DATE)
            gap_pct = _num(gap.get("gap_pct"))
            _put(ws, row, 3, gap_pct, font=_sign_font(gap_pct), fmt=FMT_PCT)
            _put(ws, row, 4, _num(gap.get("open")), fmt=FMT_NUM)
            _put(ws, row, 5, _num(gap.get("prev_close")), fmt=FMT_NUM)
            row += 1


def _ordered_weekdays(weekdays: Dict[str, Any]) -> List[str]:
    """Jours dans l'ordre de la semaine ; les inconnus sont conservés à la fin."""
    known = [d for d in WEEKDAY_ORDER if d in weekdays]
    extra = [d for d in weekdays if d not in WEEKDAY_ORDER]
    return known + extra


# --------------------------------------------------------------------------
# Feuille 3 — Notizie
# --------------------------------------------------------------------------

def _sheet_notizie(wb, news: Dict[str, Any]) -> None:
    ws = wb.create_sheet(title="Notizie")
    _set_widths(ws, [70, 22, 20, 46])

    ws["A1"] = "Titoli di stampa (nessun commento, nessuna interpretazione)"
    ws["A1"].font = TITLE_FONT

    news = news or {}
    row = 3
    _write_headers(ws, row, NEWS_HEADERS)
    row += 1
    for item in (news.get("items") or []):
        item = item or {}
        _put(ws, row, 1, _safe_text(item.get("title")), align=WRAP_ALIGN)
        _put(ws, row, 2, _safe_text(item.get("source")), align=LEFT_ALIGN)
        _put(ws, row, 3, _local_dt(item.get("published")), fmt=FMT_DATETIME)
        url = item.get("url")
        # Un lien n'est posé que s'il est http(s) : une URL vide ou exotique
        # produirait un hyperlien mort dans Excel.
        if isinstance(url, str) and url.startswith(("http://", "https://")):
            _put(ws, row, 4, url, align=LEFT_ALIGN, link=url)
        else:
            _put(ws, row, 4, None, align=LEFT_ALIGN)
        row += 1

    themes = news.get("themes") or []
    if themes:
        row += 1
        ws.cell(row=row, column=1).value = "Temi ricorrenti"
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        row += 1
        _write_headers(ws, row, ["Tema", "Occorrenze", "Esempi"])
        row += 1
        for theme in themes:
            theme = theme or {}
            _put(ws, row, 1, _safe_text(theme.get("theme")), align=LEFT_ALIGN)
            _put(ws, row, 2, _num(theme.get("count")), fmt=FMT_INT)
            examples = theme.get("examples") or []
            joined = " · ".join(str(e) for e in examples) if examples else None
            _put(ws, row, 3, _safe_text(joined), align=WRAP_ALIGN)
            row += 1

    failed = news.get("sources_failed") or []
    if failed:
        row += 1
        ws.cell(row=row, column=1).value = "Fonti non disponibili"
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        row += 1
        _write_headers(ws, row, ["Fonte non raggiungibile", "Errore"])
        row += 1
        for entry in failed:
            entry = entry or {}
            _put(ws, row, 1, _safe_text(entry.get("source")), align=LEFT_ALIGN)
            _put(ws, row, 2, _safe_text(entry.get("error")), align=LEFT_ALIGN)
            row += 1


# --------------------------------------------------------------------------
# API publique
# --------------------------------------------------------------------------

def write_workbook(path, snapshot, history=None, news=None):
    """Écrit le classeur et retourne le chemin écrit (str).

    `snapshot` : dict produit par `pulse.snapshot.build_snapshot`.
    `history`  : dict de `build_history_stats` (feuille Statistiche si fourni).
    `news`     : dict {items, themes, sources_ok, sources_failed} (feuille
                 Notizie si fourni). Toute clé peut manquer.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # retire le "Sheet" par défaut

    _sheet_mercati(wb, snapshot or {})
    if history is not None:
        _sheet_statistiche(wb, history)
    if news is not None:
        _sheet_notizie(wb, news)

    wb.properties.creator = "Market Pulse — OmenServer"
    wb.properties.lastModifiedBy = "Market Pulse"
    out = str(path)
    wb.save(out)
    return out
