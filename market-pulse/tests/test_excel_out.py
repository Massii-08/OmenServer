"""Export Excel — écrit dans tmp_path puis RELIT le classeur avec openpyxl.

Tous les tests tournent hors ligne sur les fixtures réelles capturées le
2026-07-28 (snapshot_sample.json / history_sample.json).
"""
import json
import os

import openpyxl
import pytest

from pulse.excel_out import write_workbook

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")


def _load(name):
    with open(os.path.join(FIXTURES, name)) as f:
        return json.load(f)


@pytest.fixture
def snapshot():
    return _load("snapshot_sample.json")


@pytest.fixture
def history():
    return _load("history_sample.json")


NEWS = {
    "items": [
        {
            "title": "La BCE lascia i tassi invariati",
            "source": "Il Sole 24 Ore",
            "url": "https://example.com/bce",
            "published": 1785250000,
            "lang": "it",
        },
        {
            "title": "Wall Street apre in rialzo",
            "source": "Reuters",
            "url": "",  # lien absent → aucun hyperlien à poser
            "published": None,
            "lang": "it",
        },
        {
            "title": "=SOMMA(A1:A2) titolo ostile <&> \"virgolette\"",
            "source": "@fonte-sospetta",
            "url": "https://example.com/x",
            "published": 1785250000,
            "lang": "it",
        },
    ],
    "themes": [{"theme": "Banche centrali", "count": 4, "examples": ["BCE", "Fed"]}],
    "sources_ok": ["Reuters"],
    "sources_failed": [{"source": "ANSA", "error": "timeout"}],
}


def _read(path):
    return openpyxl.load_workbook(path)


def _find_row(ws, value, column=1):
    """Numéro de ligne dont la cellule de `column` vaut `value` (None sinon)."""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=column).value == value:
            return row
    return None


def _header_row(ws, first_header):
    row = _find_row(ws, first_header)
    assert row is not None, "en-tête %r introuvable" % first_header
    return row


def _row_of_market(ws, label):
    row = _find_row(ws, label)
    assert row is not None, "marché %r introuvable" % label
    return row


def _write(tmp_path, *args, **kwargs):
    path = os.path.join(str(tmp_path), "market_pulse.xlsx")
    return write_workbook(path, *args, **kwargs)


# --------------------------------------------------------------------------
# Structure
# --------------------------------------------------------------------------

def test_returns_written_path_and_creates_file(tmp_path, snapshot):
    out = _write(tmp_path, snapshot)
    assert out.endswith("market_pulse.xlsx")
    assert os.path.isfile(out)


def test_sheets_present_when_all_sources_given(tmp_path, snapshot, history):
    wb = _read(_write(tmp_path, snapshot, history=history, news=NEWS))
    assert wb.sheetnames == ["Mercati", "Statistiche", "Notizie"]


def test_optional_sheets_absent_without_history_and_news(tmp_path, snapshot):
    wb = _read(_write(tmp_path, snapshot))
    assert wb.sheetnames == ["Mercati"]
    assert "Statistiche" not in wb.sheetnames
    assert "Notizie" not in wb.sheetnames


def test_avviso_no_investment_advice_on_first_sheet(tmp_path, snapshot):
    ws = _read(_write(tmp_path, snapshot))["Mercati"]
    texts = [
        ws.cell(row=r, column=1).value
        for r in range(1, 6)
        if isinstance(ws.cell(row=r, column=1).value, str)
    ]
    joined = " ".join(texts).lower()
    assert "nessun consiglio" in joined
    assert "informativo" in joined


# --------------------------------------------------------------------------
# Feuille Mercati
# --------------------------------------------------------------------------

def test_mercati_headers_in_italian(tmp_path, snapshot):
    ws = _read(_write(tmp_path, snapshot))["Mercati"]
    hrow = _header_row(ws, "Mercato")
    headers = [ws.cell(row=hrow, column=c).value for c in range(1, 12)]
    assert headers == [
        "Mercato", "Area", "Tipo", "Stato", "Ora locale", "Ultimo",
        "Chiusura prec.", "Var. %", "Gap apertura %", "Data gap", "Valuta",
    ]


def test_mercati_has_one_row_per_market(tmp_path, snapshot):
    ws = _read(_write(tmp_path, snapshot))["Mercati"]
    hrow = _header_row(ws, "Mercato")
    assert ws.max_row - hrow == len(snapshot["markets"]) == 20


def test_numbers_are_real_numbers_not_text(tmp_path, snapshot):
    ws = _read(_write(tmp_path, snapshot))["Mercati"]
    row = _row_of_market(ws, "FTSE MIB (Milano)")
    assert ws.cell(row=row, column=6).value == pytest.approx(51698.19)
    assert isinstance(ws.cell(row=row, column=6).value, float)
    assert ws.cell(row=row, column=7).value == pytest.approx(52055.0)
    # La variation est stockée telle que publiée (-0.69 = -0,69 %)
    assert ws.cell(row=row, column=8).value == pytest.approx(-0.69)
    assert "%" in ws.cell(row=row, column=8).number_format


def test_labels_translated_to_italian(tmp_path, snapshot):
    ws = _read(_write(tmp_path, snapshot))["Mercati"]
    row = _row_of_market(ws, "FTSE MIB (Milano)")
    assert ws.cell(row=row, column=2).value == "Europa"
    assert ws.cell(row=row, column=3).value == "Indice"
    assert ws.cell(row=row, column=4).value == "chiuso"      # clock.status closed
    assert ws.cell(row=row, column=5).value == "18:58"
    assert ws.cell(row=row, column=11).value == "EUR"
    # Un marché ouvert est signalé en majuscules
    usa = _row_of_market(ws, "S&P 500")
    assert ws.cell(row=usa, column=4).value == "APERTO"


def test_negative_variation_is_red_and_positive_is_green(tmp_path, snapshot):
    ws = _read(_write(tmp_path, snapshot))["Mercati"]
    neg = ws.cell(row=_row_of_market(ws, "Nikkei 225 (Tokyo)"), column=8)  # -3,95 %
    pos = ws.cell(row=_row_of_market(ws, "DAX (Francoforte)"), column=8)   # +0,41 %
    assert neg.value == pytest.approx(-3.95)
    assert "FF0000" in neg.font.color.rgb
    assert pos.value == pytest.approx(0.41)
    assert "FF0000" not in pos.font.color.rgb
    assert pos.font.color.rgb != neg.font.color.rgb


def test_rate_variation_expressed_in_basis_points(tmp_path, snapshot):
    """kind == 'rate' : le « prix » EST un taux → niveau en %, variation en pb."""
    ws = _read(_write(tmp_path, snapshot))["Mercati"]
    row = _row_of_market(ws, "Treasury USA 10 anni")
    price = ws.cell(row=row, column=6)
    assert price.value == pytest.approx(4.592)
    assert '"%"' in price.number_format          # niveau affiché 4,59%
    var = ws.cell(row=row, column=8)
    assert var.value == pytest.approx(-4.9)      # 4,592 - 4,641 = -4,9 pb
    assert "pb" in var.number_format
    assert "%" not in var.number_format
    gap = ws.cell(row=row, column=9)
    assert gap.value == pytest.approx(-1.1)      # 4,630 - 4,641 = -1,1 pb
    assert "pb" in gap.number_format


def test_fx_uses_four_decimals(tmp_path, snapshot):
    ws = _read(_write(tmp_path, snapshot))["Mercati"]
    row = _row_of_market(ws, "EUR/USD")
    assert ws.cell(row=row, column=6).number_format == "#,##0.0000"
    assert ws.cell(row=row, column=6).value == pytest.approx(1.14)


def test_gap_date_is_a_real_date(tmp_path, snapshot):
    ws = _read(_write(tmp_path, snapshot))["Mercati"]
    row = _row_of_market(ws, "S&P 500")
    cell = ws.cell(row=row, column=10)
    assert cell.value.strftime("%Y-%m-%d") == "2026-07-28"


# --------------------------------------------------------------------------
# Robustesse
# --------------------------------------------------------------------------

def test_zero_markets_still_writes_headers(tmp_path):
    ws = _read(_write(tmp_path, {"generated_at": 1785257927, "markets": [], "errors": []}))["Mercati"]
    hrow = _header_row(ws, "Mercato")
    assert ws.cell(row=hrow, column=6).value == "Ultimo"
    assert ws.max_row == hrow


def test_none_values_leave_cells_empty(tmp_path):
    snap = {
        "generated_at": 1785257927,
        "markets": [{
            "symbol": "X", "label": "Mercato vuoto", "region": None, "kind": None,
            "name": None, "currency": None, "price": None, "prev_close": None,
            "change_pct": None, "clock": None, "gap": None, "gap_is_today": False,
        }],
        "errors": [],
    }
    ws = _read(_write(tmp_path, snap))["Mercati"]
    row = _row_of_market(ws, "Mercato vuoto")
    for col in (5, 6, 7, 8, 9, 10, 11):
        value = ws.cell(row=row, column=col).value
        assert value is None or value == "", "col %d = %r" % (col, value)


def test_missing_snapshot_keys_do_not_crash(tmp_path):
    ws = _read(_write(tmp_path, {}))["Mercati"]
    assert _header_row(ws, "Mercato")


# --------------------------------------------------------------------------
# Feuille Statistiche
# --------------------------------------------------------------------------

def test_statistiche_weekday_values(tmp_path, snapshot, history):
    ws = _read(_write(tmp_path, snapshot, history=history))["Statistiche"]
    hrow = _header_row(ws, "Mercato")
    assert [ws.cell(row=hrow, column=c).value for c in range(1, 8)] == [
        "Mercato", "Sedute", "Giorno", "Sedute (giorno)",
        "Gap medio %", "Gap medio assoluto %", "Sedute in rialzo %",
    ]
    # Première ligne = S&P 500 / lunedì (l'ordre des jours est respecté)
    row = hrow + 1
    assert ws.cell(row=row, column=1).value == "S&P 500"
    assert ws.cell(row=row, column=2).value == 251
    assert ws.cell(row=row, column=3).value == "lunedì"
    assert ws.cell(row=row, column=4).value == 48
    assert ws.cell(row=row, column=5).value == pytest.approx(0.184)
    assert ws.cell(row=row, column=6).value == pytest.approx(0.391)
    assert ws.cell(row=row, column=7).value == pytest.approx(58.3)
    assert ws.cell(row=row + 1, column=3).value == "martedì"


def test_statistiche_biggest_gaps_block(tmp_path, snapshot, history):
    ws = _read(_write(tmp_path, snapshot, history=history))["Statistiche"]
    # Le bloc se repère par son sous-titre : « Mercato » est l'en-tête des DEUX
    # blocs de la feuille, et « Data » vit en colonne 2, pas en colonne 1 —
    # chercher « Data » en colonne 1 ne pouvait rien trouver.
    subtitle = _find_row(ws, "Maggiori gap registrati")
    assert subtitle is not None, "sous-titre du bloc des plus gros gaps introuvable"
    hrow = subtitle + 1
    assert [ws.cell(row=hrow, column=c).value for c in range(1, 5)] == [
        "Mercato", "Data", "Gap %", "Apertura",
    ]
    row = hrow + 1
    assert ws.cell(row=row, column=1).value == "S&P 500"
    assert ws.cell(row=row, column=2).value.strftime("%Y-%m-%d") == "2026-04-08"
    assert ws.cell(row=row, column=3).value == pytest.approx(2.08)
    assert ws.cell(row=row, column=4).value == pytest.approx(6754.35986328125)


def test_statistiche_tolerates_empty_stats(tmp_path, snapshot):
    wb = _read(_write(tmp_path, snapshot, history={"stats": {}, "errors": []}))
    assert "Statistiche" in wb.sheetnames


# --------------------------------------------------------------------------
# Feuille Notizie
# --------------------------------------------------------------------------

def test_notizie_headers_and_hyperlink(tmp_path, snapshot):
    ws = _read(_write(tmp_path, snapshot, news=NEWS))["Notizie"]
    hrow = _header_row(ws, "Titolo")
    assert [ws.cell(row=hrow, column=c).value for c in range(1, 5)] == [
        "Titolo", "Fonte", "Ora", "Link",
    ]
    first = ws.cell(row=hrow + 1, column=1)
    assert first.value == "La BCE lascia i tassi invariati"
    assert ws.cell(row=hrow + 1, column=2).value == "Il Sole 24 Ore"
    link = ws.cell(row=hrow + 1, column=4)
    assert link.hyperlink is not None
    assert link.hyperlink.target == "https://example.com/bce"


def test_notizie_empty_url_leaves_no_broken_hyperlink(tmp_path, snapshot):
    ws = _read(_write(tmp_path, snapshot, news=NEWS))["Notizie"]
    hrow = _header_row(ws, "Titolo")
    row = hrow + 2  # 2e item : url vide, published None
    assert ws.cell(row=row, column=1).value == "Wall Street apre in rialzo"
    assert ws.cell(row=row, column=4).hyperlink is None
    assert ws.cell(row=row, column=3).value in (None, "")


def test_formula_injection_is_neutralised(tmp_path, snapshot):
    ws = _read(_write(tmp_path, snapshot, news=NEWS))["Notizie"]
    hrow = _header_row(ws, "Titolo")
    hostile = ws.cell(row=hrow + 3, column=1)
    assert hostile.value.startswith("'=SOMMA(A1:A2)")
    assert hostile.data_type != "f"
    assert "virgolette" in hostile.value
    assert ws.cell(row=hrow + 3, column=2).value.startswith("'@")


def test_notizie_themes_and_failed_sources(tmp_path, snapshot):
    ws = _read(_write(tmp_path, snapshot, news=NEWS))["Notizie"]
    trow = _header_row(ws, "Tema")
    assert ws.cell(row=trow + 1, column=1).value == "Banche centrali"
    assert ws.cell(row=trow + 1, column=2).value == 4
    assert "BCE" in ws.cell(row=trow + 1, column=3).value
    frow = _header_row(ws, "Fonte non raggiungibile")
    assert ws.cell(row=frow + 1, column=1).value == "ANSA"
    assert ws.cell(row=frow + 1, column=2).value == "timeout"


def test_notizie_tolerates_partial_payload(tmp_path, snapshot):
    """`news` sans 'themes' ni 'sources_failed', items incomplets."""
    partial = {"items": [{"title": "Solo titolo"}]}
    ws = _read(_write(tmp_path, snapshot, news=partial))["Notizie"]
    hrow = _header_row(ws, "Titolo")
    assert ws.cell(row=hrow + 1, column=1).value == "Solo titolo"
    assert ws.cell(row=hrow + 1, column=2).value in (None, "")
    assert _find_row(ws, "Tema") is None
