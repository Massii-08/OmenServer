"""
Processore Excel per la lista delle obbligazioni.

Gestisce la lettura e scrittura del file "Lista acquisti-2026.xlsx".
Supporta i 4 fogli: Euro, USD, GBP, Vale.
"""

import logging
import os
import shutil
from datetime import date, datetime
from typing import List, Dict, Optional, Tuple

import re

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from scraper.models import BondData

logger = logging.getLogger(__name__)

# Sfondo blu per le righe con errori da controllare
BLUE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)  # Per rimuovere il blu

# Fonts pour la coloration prix (seuil configurable, défaut 101)
RED_FONT = Font(color="FF0000")    # Prix < seuil → rouge (bonne affaire)
BLACK_FONT = Font(color="000000")  # Prix >= seuil → noir (au-dessus du pair)

# Seuil par défaut pour la coloration
DEFAULT_PRICE_THRESHOLD = 101

# Alignement centré (comme les lignes existantes)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

# Mois italiens pour le format date (es: "mag.15", "feb.23")
ITALIAN_MONTHS = {
    1: 'gen', 2: 'feb', 3: 'mar', 4: 'apr', 5: 'mag', 6: 'giu',
    7: 'lug', 8: 'ago', 9: 'set', 10: 'ott', 11: 'nov', 12: 'dic',
}


def _format_date_italian(d) -> str:
    """Formate une date en format italien: 'mag.15', 'feb.23', etc."""
    if isinstance(d, str):
        # Tenter de parser DD.MM.YYYY ou YYYY-MM-DD
        for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
            try:
                d = datetime.strptime(d, fmt).date()
                break
            except ValueError:
                continue
        else:
            return d  # Retourner tel quel si non parsable
    if isinstance(d, (date, datetime)):
        month_abbr = ITALIAN_MONTHS.get(d.month, str(d.month))
        year_short = str(d.year)[-2:]  # 2025 → "25"
        return f"{month_abbr}.{year_short}"
    return str(d)


def _format_as_k(value) -> str:
    """Formate un nombre en notation 'Xk': 2000→'2k', 500000000→'500,000k'."""
    try:
        num = float(value)
        k_value = int(num / 1000)
        if k_value == 0:
            return str(int(num))  # Nombres < 1000 : afficher tel quel
        # Formater avec séparateur de milliers + 'k'
        return f"{k_value:,}k"
    except (ValueError, TypeError):
        return str(value)


# Mappatura colonne per i fogli Euro, USD, GBP (stessa struttura)
STANDARD_COLUMNS = {
    'papy': 'A',        # Col A: PAPY (segno se acquistato)
    'name': 'B',        # Col B: Nome bond
    'emission': 'C',    # Col C: Data emissione
    'isin': 'D',        # Col D: ISIN
    'price': 'E',       # Col E: Prezzo
    'yield': 'F',       # Col F: Yield
    'rating': 'G',      # Col G: Rating
    'volume': 'H',      # Col H: Volume
    'min_piece': 'I',   # Col I: Pezzo minimo
    'header_row': 3,    # Riga dell'header
    'data_start': 4,    # Prima riga dati
}

# Mappatura colonne per il foglio Vale (struttura diversa)
VALE_COLUMNS = {
    'papy': 'A',        # Col A: PAPY
    'name': 'B',        # Col B: Nome bond
    'emission': 'C',    # Col C: Data emissione
    'isin': 'D',        # Col D: ISIN
    'price': 'E',       # Col E: Prezzo
    'currency': 'F',    # Col F: Valuta (EUR/USD)
    'yield': 'G',       # Col G: Yield
    'rating': 'H',      # Col H: Rating
    'volume': 'I',      # Col I: Volume
    'min_piece': 'J',   # Col J: Pezzo minimo
    'header_row': 2,    # Riga dell'header
    'data_start': 3,    # Prima riga dati
}


class BondExcelProcessor:
    """
    Processore per leggere e scrivere dati nel file Excel delle obbligazioni.
    """
    
    def __init__(self, filepath: str, price_threshold: float = None):
        """
        Args:
            filepath: Percorso al file Excel (Lista acquisti-2026.xlsx)
            price_threshold: Seuil de prix pour coloration rouge/noir (défaut: 101)
        """
        self.filepath = filepath
        self.price_threshold = price_threshold if price_threshold is not None else DEFAULT_PRICE_THRESHOLD
        self.wb = None
        self._load()
    
    def _load(self):
        """Carica il workbook Excel."""
        if not os.path.exists(self.filepath):
            raise FileNotFoundError(f"File non trovato: {self.filepath}")
        
        self.wb = openpyxl.load_workbook(self.filepath)
        logger.info(f"File caricato: {self.filepath}")
        logger.info(f"Fogli: {self.wb.sheetnames}")
    
    def _get_columns(self, sheet_name: str) -> dict:
        """Restituisce la mappatura colonne per un foglio."""
        if sheet_name == 'Vale':
            return VALE_COLUMNS
        return STANDARD_COLUMNS
    
    def get_all_bonds(self) -> List[Dict]:
        """
        Estrae tutte le obbligazioni da tutti i fogli.
        
        Returns:
            Lista di dizionari con i dati di ogni bond
        """
        all_bonds = []
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            cols = self._get_columns(sheet_name)
            
            for row in range(cols['data_start'], ws.max_row + 1):
                isin_cell = ws[f"{cols['isin']}{row}"]
                name_cell = ws[f"{cols['name']}{row}"]
                
                # Salta righe vuote o senza ISIN
                if not isin_cell.value:
                    continue
                
                isin = str(isin_cell.value).strip()
                
                # Verifica che sia un ISIN valido (inizia con 2 lettere + 10 caratteri)
                if len(isin) < 12:
                    continue
                
                bond_info = {
                    'sheet': sheet_name,
                    'row': row,
                    'isin': isin,
                    'name': str(name_cell.value).strip() if name_cell.value else '',
                    'price': ws[f"{cols['price']}{row}"].value,
                    'yield': ws[f"{cols['yield']}{row}"].value,
                    'rating': ws[f"{cols['rating']}{row}"].value,
                    'emission': ws[f"{cols['emission']}{row}"].value,
                }
                
                # Per il foglio Vale, aggiungi la valuta
                if sheet_name == 'Vale':
                    bond_info['currency'] = ws[f"{cols['currency']}{row}"].value
                else:
                    # La valuta è nel nome del foglio / header
                    currency_map = {'Euro': 'EUR', 'USD': 'USD', 'GBP': 'GBP'}
                    bond_info['currency'] = currency_map.get(sheet_name, 'EUR')
                
                all_bonds.append(bond_info)
        
        logger.info(f"Trovate {len(all_bonds)} obbligazioni totali")
        return all_bonds
    
    def get_bonds_by_sheet(self, sheet_name: str) -> List[Dict]:
        """Restituisce le obbligazioni di un singolo foglio."""
        all_bonds = self.get_all_bonds()
        return [b for b in all_bonds if b['sheet'] == sheet_name]
    
    def update_yield(self, sheet_name: str, row: int, new_yield: float):
        """
        Aggiorna il yield di un'obbligazione specifica.
        Formato: percentuale arrotondata a 2 decimali (es: 4.35%)
        """
        ws = self.wb[sheet_name]
        cols = self._get_columns(sheet_name)
        
        cell = ws[f"{cols['yield']}{row}"]
        old_value = cell.value
        # Arrotondare a 2 cifre decimali (es: 0.043527 → 0.0435)
        cell.value = round(new_yield, 4)
        
        # Formatta come percentuale a 2 decimali (es: 4.35%)
        cell.number_format = '0.00%'
        cell.alignment = CENTER_ALIGN
        
        logger.info(
            f"  Yield aggiornato: {sheet_name} riga {row}: "
            f"{old_value} → {new_yield:.2%}"
        )
    
    def update_price(self, sheet_name: str, row: int, new_price: float):
        """
        Aggiorna il prezzo di un'obbligazione.
        
        Args:
            sheet_name: Nome del foglio
            row: Numero di riga
            new_price: Nuovo prezzo
        """
        ws = self.wb[sheet_name]
        cols = self._get_columns(sheet_name)
        
        cell = ws[f"{cols['price']}{row}"]
        old_value = cell.value
        cell.value = new_price
        cell.alignment = CENTER_ALIGN
        
        logger.info(
            f"  Prezzo aggiornato: {sheet_name} riga {row}: "
            f"{old_value} → {new_price}"
        )
    
    def update_rating(self, sheet_name: str, row: int, rating: str):
        """Aggiorna il rating di un'obbligazione."""
        ws = self.wb[sheet_name]
        cols = self._get_columns(sheet_name)
        
        cell = ws[f"{cols['rating']}{row}"]
        if cell.value in (None, '?', '? '):
            cell.value = rating
            cell.alignment = CENTER_ALIGN
            logger.info(f"  Rating aggiornato: {sheet_name} riga {row}: {rating}")
    
    def update_bond_full(self, sheet_name: str, row: int, bond_data: BondData):
        """
        Aggiorna tutti i campi disponibili di un'obbligazione.
        Remplit les champs vides avec les données scrapées (sauf rating).
        
        Args:
            sheet_name: Nome del foglio
            row: Numero di riga
            bond_data: Dati completi del bond
        """
        if bond_data.current_price is not None:
            self.update_price(sheet_name, row, bond_data.current_price)
        
        if bond_data.calculated_yield is not None:
            self.update_yield(sheet_name, row, bond_data.calculated_yield)
        
        if bond_data.rating and bond_data.rating != '?':
            self.update_rating(sheet_name, row, bond_data.rating)
        
        # Remplir les champs vides avec les données de la bourse
        self.fill_empty_fields(sheet_name, row, bond_data)
        
        # Corriger le nom si les données scrapées fournissent les infos
        self.update_name(sheet_name, row, bond_data)
        
        # Appliquer la coloration rouge/noir selon le prix
        if bond_data.current_price is not None:
            self.apply_price_color(sheet_name, row, bond_data.current_price)
    
    def fill_empty_fields(self, sheet_name: str, row: int, bond_data: BondData):
        """
        Remplit les cellules vides avec les données trouvées sur Deutsche Börse.
        Ne touche PAS au rating (pas disponible sur la bourse).
        Formate les valeurs comme les lignes existantes :
          - Date : format italien (mag.15, feb.23)
          - Volume : notation Xk (500,000k)
          - Min Piece : notation Xk (2k, 10k)
          - Alignement centré
        """
        ws = self.wb[sheet_name]
        cols = self._get_columns(sheet_name)
        
        # --- Emission (date d'émission) — format italien : "mag.15" ---
        emission_cell = ws[f"{cols['emission']}{row}"]
        if emission_cell.value is None and bond_data.issue_date is not None:
            formatted_date = _format_date_italian(bond_data.issue_date)
            emission_cell.value = formatted_date
            emission_cell.alignment = CENTER_ALIGN
            logger.info(f"  📝 Emission remplie: {sheet_name}:{row} → {formatted_date}")
        
        # --- Volume — format Xk : "500,000k" ---
        volume_cell = ws[f"{cols['volume']}{row}"]
        if volume_cell.value is None and bond_data.volume is not None:
            formatted_vol = _format_as_k(bond_data.volume)
            volume_cell.value = formatted_vol
            volume_cell.alignment = CENTER_ALIGN
            logger.info(f"  📝 Volume rempli: {sheet_name}:{row} → {formatted_vol}")
        
        # --- Min Piece — format Xk : "2k", "10k" ---
        min_piece_cell = ws[f"{cols['min_piece']}{row}"]
        if min_piece_cell.value is None and bond_data.min_piece is not None:
            formatted_mp = _format_as_k(bond_data.min_piece)
            min_piece_cell.value = formatted_mp
            min_piece_cell.alignment = CENTER_ALIGN
            logger.info(f"  📝 Min piece rempli: {sheet_name}:{row} → {formatted_mp}")
        
        # --- Valuta (seulement pour le foglio Vale) ---
        if sheet_name == 'Vale' and 'currency' in cols:
            currency_cell = ws[f"{cols['currency']}{row}"]
            if currency_cell.value is None and bond_data.currency:
                currency_cell.value = bond_data.currency
                currency_cell.alignment = CENTER_ALIGN
                logger.info(f"  📝 Valuta remplie: {sheet_name}:{row} → {bond_data.currency}")
    
    def update_name(self, sheet_name: str, row: int, bond_data: BondData):
        """
        Corrige le nom du bond pour qu'il corresponde au format standard:
        "ISSUER COUPON% - DD.MM.YY" ou "ISSUER COUPON - DD.MM.YY"
        
        Utilise les données scrapées (nom, coupon, scadenza) pour reconstruire
        un nom propre si le nom actuel est incorrect ou incomplet.
        """
        ws = self.wb[sheet_name]
        cols = self._get_columns(sheet_name)
        name_cell = ws[f"{cols['name']}{row}"]
        current_name = str(name_cell.value).strip() if name_cell.value else ''
        
        # Si pas de données scrapées, on ne peut rien corriger
        if not bond_data.name:
            return
        
        # Construire le nom formaté à partir des données de la bourse
        formatted_name = self._format_bond_name(
            scraped_name=bond_data.name,
            coupon_rate=bond_data.coupon_rate,
            maturity_date=bond_data.maturity_date,
        )
        
        if not formatted_name:
            return
        
        # Comparer : corriger si le nom actuel est vide ou significativement différent
        if not current_name or self._name_needs_correction(current_name, formatted_name):
            old_name = current_name or '(vide)'
            name_cell.value = formatted_name
            logger.info(f"  ✏️  Nom corrigé: {sheet_name}:{row}")
            logger.info(f"     Ancien: {old_name}")
            logger.info(f"     Nouveau: {formatted_name}")
    
    def _format_bond_name(self, scraped_name: str, coupon_rate: float = None, 
                          maturity_date=None) -> str:
        """
        Formate un nom de bond selon la convention du fichier Excel:
        "ISSUER COUPON% - DD.MM.YY" ou "ISSUER COUPON,XXX - DD.MM.YY"
        
        Exemples de noms dans le fichier:
            Deutsche Post AG 3,375% - 03.07.33
            Dell 5,3 - 01.10.29
            POWER FIN  1,841 - 21.09.28
        """
        if not scraped_name:
            return ''
        
        # Extraire le nom de l'émetteur depuis le nom scrapé
        # Le nom scrapé de Deutsche Börse est souvent complet, on extrait l'émetteur
        issuer = scraped_name
        
        # Supprimer les parties connues du nom scrapé (coupon, dates, etc.)
        # Les noms DB sont souvent "ISSUER COUPON DATE" en un seul bloc
        # On nettoie les patterns courants
        issuer = re.sub(r'\d{1,2}[./]\d{1,2}[./]\d{2,4}', '', issuer)  # Dates
        issuer = re.sub(r'\d+[,.]\d*\s*%', '', issuer)                   # Pourcentages
        issuer = re.sub(r'\s*[-–]\s*$', '', issuer)                      # Tiret final
        issuer = re.sub(r'\s+', ' ', issuer).strip()                     # Espaces multiples
        
        # Construire le nom formaté
        parts = [issuer]
        
        if coupon_rate is not None:
            # Formater le coupon : virgule pour décimales (convention européenne)
            coupon_str = f"{coupon_rate}".replace('.', ',')
            # Supprimer les zéros inutiles : "3,0" → "3", "3,375" reste
            if ',' in coupon_str:
                coupon_str = coupon_str.rstrip('0').rstrip(',')
            parts.append(coupon_str)
        
        if maturity_date is not None:
            # Format DD.MM.YY (2 chiffres année)
            date_str = maturity_date.strftime('%d.%m.%y')
            parts.append(f"- {date_str}")
        
        return ' '.join(parts)
    
    def _name_needs_correction(self, current_name: str, formatted_name: str) -> bool:
        """
        Détermine si le nom actuel doit être corrigé.
        Retourne True si le nom est significativement différent du format attendu.
        """
        # Normaliser pour comparaison
        current_lower = current_name.lower().strip()
        formatted_lower = formatted_name.lower().strip()
        
        # Si identiques, pas de correction
        if current_lower == formatted_lower:
            return False
        
        # Vérifier si le nom actuel a le bon format de base: 
        # doit contenir un nombre (coupon) et une date (DD.MM.YY)
        has_coupon = bool(re.search(r'\d+[,.]\d+|\d+\s*%', current_name))
        has_date = bool(re.search(r'\d{1,2}[./]\d{1,2}[./]\d{2,4}', current_name))
        has_dash = bool(re.search(r'\s*[-–]\s*', current_name))
        
        # Si le format de base est correct (coupon + tiret + date), ne pas toucher
        if has_coupon and has_date and has_dash:
            return False
        
        # Sinon, le nom a besoin de correction
        return True
    
    def apply_price_color(self, sheet_name: str, row: int, price: float):
        """
        Applique la coloration des caractères selon le prix:
        - Prix < seuil → rouge (obligation en-dessous du pair, bonne opportunité)
        - Prix >= seuil → noir (au-dessus du pair)
        
        Le seuil est configurable via self.price_threshold (défaut: 101).
        """
        ws = self.wb[sheet_name]
        cols = self._get_columns(sheet_name)
        threshold = self.price_threshold
        
        font = RED_FONT if price > threshold else BLACK_FONT
        
        # Appliquer à toutes les colonnes de la ligne (de A à la dernière colonne utilisée)
        all_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        if sheet_name == 'Vale':
            all_cols.append('J')  # Vale a une colonne de plus
        
        for col in all_cols:
            cell = ws[f"{col}{row}"]
            # Préserver le style existant (gras, taille) mais changer la couleur
            old_font = cell.font
            cell.font = Font(
                name=old_font.name,
                size=old_font.size,
                bold=old_font.bold,
                italic=old_font.italic,
                underline=old_font.underline,
                strike=old_font.strike,
                color=font.color,
            )
        
        color_name = 'rouge' if price > threshold else 'noir'
        logger.info(f"  🎨 Couleur {color_name}: {sheet_name}:{row} (prix={price}, seuil={threshold})")
    
    def mark_orange_dot(self, sheet_name: str, row: int):
        """
        Pallino arancione (●) → bond sauté (données insuffisantes).
        Colonne J (ou K pour Vale).
        """
        ws = self.wb[sheet_name]
        dot_col = 'K' if sheet_name == 'Vale' else 'J'
        cell = ws[f"{dot_col}{row}"]
        cell.value = "●"
        cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        cell.font = Font(color="FF8C00")
        cell.alignment = CENTER_ALIGN
        logger.info(f"  🟠 {sheet_name}:{row} pallino arancione (skip)")
    
    def mark_red_dot(self, sheet_name: str, row: int):
        """
        Pallino rosso (●) → erreur (ISIN non trouvé, scraping échoué).
        Colonne J (ou K pour Vale).
        """
        ws = self.wb[sheet_name]
        dot_col = 'K' if sheet_name == 'Vale' else 'J'
        cell = ws[f"{dot_col}{row}"]
        cell.value = "●"
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        cell.font = Font(color="FF0000")
        cell.alignment = CENTER_ALIGN
        logger.info(f"  🔴 {sheet_name}:{row} pallino rosso (erreur)")
    
    def clear_dot(self, sheet_name: str, row: int):
        """
        Enlève le pallino (orange ou rouge) quand tout est OK.
        """
        ws = self.wb[sheet_name]
        dot_col = 'K' if sheet_name == 'Vale' else 'J'
        cell = ws[f"{dot_col}{row}"]
        cell.value = None
        cell.fill = NO_FILL
        logger.info(f"  ✅ {sheet_name}:{row} pallino enlevé (tout OK)")
    
    # Rétrocompatibilité
    def mark_blue_dot(self, sheet_name: str, row: int):
        self.mark_orange_dot(sheet_name, row)
    def apply_blue_row(self, sheet_name: str, row: int):
        self.mark_red_dot(sheet_name, row)
    def clear_blue_row(self, sheet_name: str, row: int):
        self.clear_dot(sheet_name, row)
    
    def save(self, backup: bool = True) -> str:
        """
        Salva in un NUOVO file e elimina la copia interna.
        
        Returns:
            Percorso del nuovo file salvato
        """
        # Genera nome output: Lista acquisti-2026_AGGIORNATO.xlsx
        base = os.path.splitext(self.filepath)[0]
        output_path = f"{base}_AGGIORNATO.xlsx"
        
        # Rimuovi metadati autore
        self.wb.properties.creator = ''
        self.wb.properties.lastModifiedBy = ''
        
        self.wb.save(output_path)
        logger.info(f"✅ Nuovo file salvato: {output_path}")
        logger.info(f"📂 File originale NON modificato: {self.filepath}")
        
        return output_path
    
    def get_summary(self) -> str:
        """Restituisce un riepilogo del contenuto del file."""
        all_bonds = self.get_all_bonds()
        
        summary_lines = [
            f"📊 File: {os.path.basename(self.filepath)}",
            f"📑 Fogli: {', '.join(self.wb.sheetnames)}",
            f"📈 Obbligazioni totali: {len(all_bonds)}",
            "",
        ]
        
        for sheet_name in self.wb.sheetnames:
            sheet_bonds = [b for b in all_bonds if b['sheet'] == sheet_name]
            with_yield = sum(1 for b in sheet_bonds if b['yield'] is not None)
            without_yield = len(sheet_bonds) - with_yield
            
            summary_lines.append(
                f"  {sheet_name}: {len(sheet_bonds)} bond "
                f"({with_yield} con yield, {without_yield} senza)"
            )
        
        return '\n'.join(summary_lines)
